from docx import Document
import re
import pandas as pd
from openpyxl import load_workbook

# List of words to ignore capitalization (prepositions, conjunctions, etc.)
small_words = {"a", "an", "and", "but", "for", "nor", "of", "on", "or", "so", "the", "to", "up", "yet", "with", "as", "at", "by", "from", "in", "into", "near", "on", "out", "over", "through", "under", "with"}

def capitalize_product_name(product_name):
    # Split the product name into words and capitalize appropriately
    words = product_name.split()
    capitalized_words = [
        word if word.lower() in small_words else word.capitalize()
        for word in words
    ]
    return " ".join(capitalized_words)

def remove_all_hyphens(product_name):
    # Remove all hyphens from the product name
    return product_name.replace('-', '')

def add_hyphen_before_gaotek(product_name):
    # Add a hyphen before the last "GAOTek" word with spaces around it
    return re.sub(r'(gaotek)(?!.*gaotek)', r' - GAOTek', product_name, flags=re.IGNORECASE)

def remove_extra_spaces(product_name):
    # Replace multiple spaces with a single space
    return re.sub(r'\s+', ' ', product_name).strip()

def extract_product_details(word_file):
    # Load the Word document
    doc = Document(word_file)
    
    # Combine all paragraphs into a single block of text
    full_text = "\n".join([para.text for para in doc.paragraphs])
    
    # Updated pattern to capture all relevant fields in order
    pattern = r'Product Link:\s*(.*?)\s*Supplier\'s Link:\s*(.*?)\s*Price:\s*(.*?)\s*Category:\s*(.*?)\s*Product name:\s*(.*?)\s*Product ID:\s*(.*?)\s*Meta Description:\s*(.*?)\s*Technical Specifications:'
    
    # List to store extracted product details
    product_details = []

    # Search for all matches in the combined text
    matches = re.findall(pattern, full_text, re.DOTALL)  # re.DOTALL allows dot to match newline characters
    for match in matches:
        # Extract details from the match
        product_link, supplier_link, price, category, product_name, product_id, meta_description = match
        
        # Capitalize product name
        capitalized_name = capitalize_product_name(product_name.strip())
        
        # Remove all hyphens and then add one before the last GAOTek word
        no_hyphens_name = remove_all_hyphens(capitalized_name)
        final_name = add_hyphen_before_gaotek(no_hyphens_name)
        
        # Remove extra spaces
        final_name = remove_extra_spaces(final_name)

        # Prepare the final output for this product
        product_info = {
            "Product Name": final_name,
            "Product ID": product_id.strip(),
            "Category": category.strip(),
            "Product Link": product_link.strip(),
            "Supplier Link": supplier_link.strip(),
            "Meta Description": meta_description.strip()
        }
        
        # Add to the list of product details
        product_details.append(product_info)

    # Return the list of product details
    return product_details

def save_to_excel(product_details, output_file):
    # Create a pandas DataFrame from the product details list
    df = pd.DataFrame(product_details)
    
    # Save the DataFrame to an Excel file
    df.to_excel(output_file, index=False)

    # Now, open the saved Excel file and modify links to be hyperlinks
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Add hyperlinks to the Product Link and Supplier Link columns
    for row in range(2, len(df) + 2):  # Starting from row 2 (Excel is 1-indexed)
        product_link = ws[f"D{row}"].value  # Product Link is in column D
        supplier_link = ws[f"E{row}"].value  # Supplier Link is in column E
        
        # Set the cell value to be a hyperlink
        if product_link:
            ws[f"D{row}"].hyperlink = product_link
            ws[f"D{row}"].value = product_link  # This will make the link text the URL itself
        if supplier_link:
            ws[f"E{row}"].hyperlink = supplier_link
            ws[f"E{row}"].value = supplier_link  # This will make the link text the URL itself

    # Save the modified workbook
    wb.save(output_file)
    print(f"Product details have been saved to {output_file} with hyperlinks.")

# Usage example
product_details_list = extract_product_details('input.docx')
save_to_excel(product_details_list, 'output.xlsx')
